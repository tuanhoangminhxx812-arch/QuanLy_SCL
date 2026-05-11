"""
Module tự động sinh báo cáo HangMuc.xlsx (34 cột) từ các nguồn dữ liệu.
"""
import pandas as pd
import os
from data_helpers import load_tonghop, load_pm092
from form_module import (
    load_db_data, load_chitiet_by_ma, load_hopdong_list
)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
HANGMUC_OUTPUT = os.path.join(BASE_DIR, 'HangMuc.xlsx')

# 34 column names matching the original HangMuc.xls structure
HANGMUC_COLS = [
    'STT', 'Mã công trình', 'Đơn vị', 'Tên công trình', 'Nội dung',
    'Mã TS', 'Tiến độ thực hiện', 'Giá trị khái toán',
    'PAKT-DT Số QĐ ngày duyệt', 'PAKT-DT Giá trị DT',
    'Giá trị vốn KH',
    'KH ĐT Gói XL Số QĐ', 'KH ĐT Gói XL GT', 'KH ĐT Gói TB Số QĐ', 'KH ĐT Gói TB GT',
    'KQ ĐT Gói XL Số QĐ', 'KQ ĐT Gói XL GT', 'KQ ĐT Gói XL Số HĐ', 'KQ ĐT Gói XL GT HĐ',
    'KQ ĐT Gói TB Số QĐ', 'KQ ĐT Gói TB GT', 'KQ ĐT Gói TB Số HĐ', 'KQ ĐT Gói TB GT HĐ',
    'Đơn vị thi công', 'Ngày khởi công',
    'VT TCty cấp', 'VT ĐV cấp',
    'KL thực hiện GT', 'KL thực hiện %',
    'KL thanh toán GT', 'KL thanh toán %',
    'Ngày nghiệm thu', 'Giá trị quyết toán', 'Ghi chú'
]


def _safe_val(v, default=''):
    if pd.isna(v) or v is None:
        return default
    return v

def _safe_int(v):
    try: return int(float(v)) if pd.notna(v) and v != '' else 0
    except: return 0

def _to_dong(val):
    """Giữ nguyên giá trị đồng (for HangMuc report)."""
    v = _safe_int(val)
    if v == 0: return 0
    return v


def generate_hangmuc():
    """Tổng hợp dữ liệu từ tất cả nguồn → tạo HangMuc.xlsx"""
    df_th = load_tonghop()
    pm_data = load_pm092()
    db_df = load_db_data()
    
    if df_th.empty:
        return None
    
    # Merge PM_092
    if 'Mã CT' in df_th.columns:
        df_th['Thực hiện PM'] = df_th['Mã CT'].map(lambda x: pm_data.get(str(x).strip(), 0))
        if 'Thực hiện' not in df_th.columns:
            df_th['Thực hiện'] = df_th['Thực hiện PM']
        else:
            df_th['Thực hiện'] = df_th.apply(
                lambda r: r['Thực hiện PM'] if r['Thực hiện PM'] > 0 else r['Thực hiện'], axis=1
            )
    
    CAT_MAP = {
        'Đại tu lưới trung hạ thế, trạm hạ thế': 'Lưới và trạm điện',
        'Đại tu thiết bị trung hạ thế (máy phát, MBT, Re,...)': 'Thiết bị',
        'Đại tu công trình xây dựng dân dụng': 'Kiến trúc',
        'Đại tu công xa': 'Công xa',
        'Đại tu TSCĐ khác': 'Sửa chữa khác'
    }

    rows = []
    stt_counter = 0
    
    for _, row in df_th.iterrows():
        ma_ct = str(row.get('Mã CT', '')).strip()
        if not ma_ct:
            continue
        stt_counter += 1
        
        ten_ct = _safe_val(row.get('Tên công trình', ''))
        khai_toan = _safe_int(row.get('Khái toán', 0))
        thuc_hien = _safe_int(row.get('Thực hiện', 0))
        
        # Load chi tiết
        pakt_df = load_chitiet_by_ma('pakt_dt', ma_ct)
        kh_df = load_chitiet_by_ma('kh_dau_thau', ma_ct)
        kq_df = load_chitiet_by_ma('kq_dau_thau', ma_ct)
        hd_df = load_hopdong_list(ma_ct)
        vt_df = load_chitiet_by_ma('vat_tu', ma_ct)
        nt_df = load_chitiet_by_ma('nghiem_thu_qt', ma_ct)
        
        # Load from database_cong_trinh.xlsx
        db_match = db_df[db_df['Mã CT'].astype(str).str.strip() == ma_ct] if not db_df.empty and 'Mã CT' in db_df.columns else pd.DataFrame()
        ke_hoach = _safe_int(db_match.iloc[0].get('Kế hoạch', 0)) if not db_match.empty else 0
        
        # PAKT-DT
        pakt_soqd = ''
        pakt_gt = 0
        if not pakt_df.empty:
            r = pakt_df.iloc[0]
            ngay = r.get('Ngày phê duyệt', '')
            soqd = _safe_val(r.get('Số QĐ phê duyệt', ''))
            if pd.notna(ngay) and soqd:
                try:
                    if isinstance(ngay, pd.Timestamp):
                        ngay = ngay.strftime('%d/%m/%Y')
                    pakt_soqd = f"{soqd}, ngày {ngay}"
                except:
                    pakt_soqd = str(soqd)
            else:
                pakt_soqd = str(soqd)
            pakt_gt = _safe_int(r.get('Giá trị dự toán', 0))
        
        # KH đấu thầu
        def _get_kh(loai):
            sub = kh_df[kh_df['Loại gói'] == loai] if not kh_df.empty and 'Loại gói' in kh_df.columns else pd.DataFrame()
            if sub.empty: return '', 0
            r = sub.iloc[0]
            soqd = _safe_val(r.get('Số QĐ phê duyệt KH', ''))
            gt = _safe_int(r.get('GT gói thầu', 0))
            return soqd, gt
        
        kh_xl_soqd, kh_xl_gt = _get_kh('XL')
        kh_tb_soqd, kh_tb_gt = _get_kh('TB')
        
        # KQ đấu thầu
        def _get_kq(loai):
            sub = kq_df[kq_df['Loại gói'] == loai] if not kq_df.empty and 'Loại gói' in kq_df.columns else pd.DataFrame()
            if sub.empty: return '', 0
            r = sub.iloc[0]
            soqd = _safe_val(r.get('Số QĐ phê duyệt KQ', ''))
            gt = _safe_int(r.get('GT gói thầu trúng', 0))
            return soqd, gt
        
        kq_xl_soqd, kq_xl_gt = _get_kq('XL')
        kq_tb_soqd, kq_tb_gt = _get_kq('TB')
        
        # Hợp đồng - tách XL và TB
        def _get_hd(loai_filter):
            if hd_df.empty or 'Loại HĐ' not in hd_df.columns:
                return '', 0
            sub = hd_df[hd_df['Loại HĐ'] == loai_filter]
            if sub.empty: return '', 0
            r = sub.iloc[0]
            so_hd = _safe_val(r.get('Số hợp đồng', ''))
            ngay_ky = r.get('Ngày ký HĐ', '')
            if pd.notna(ngay_ky) and so_hd:
                try:
                    if isinstance(ngay_ky, pd.Timestamp):
                        ngay_ky = ngay_ky.strftime('%d/%m/%Y')
                    so_hd = f"{so_hd}, ngày {ngay_ky}"
                except: pass
            gt_hd = _safe_int(r.get('Giá trị HĐ', 0))
            return so_hd, gt_hd
        
        hd_xl_so, hd_xl_gt = _get_hd('Xây lắp')
        hd_tb_so, hd_tb_gt = _get_hd('Thiết bị')
        
        # Đơn vị thi công - from first contract
        don_vi_tc = ''
        if not hd_df.empty and 'Tên nhà thầu' in hd_df.columns:
            nha_thau = hd_df.iloc[0].get('Tên nhà thầu', '')
            don_vi_tc = _safe_val(nha_thau)
        
        # Ngày khởi công
        ngay_kc = ''
        if not db_match.empty:
            nkc = db_match.iloc[0].get('Ngày khởi công', '')
            if pd.notna(nkc):
                try:
                    if isinstance(nkc, pd.Timestamp):
                        ngay_kc = nkc.strftime('%d/%m/%Y')
                    else:
                        ngay_kc = str(nkc)
                except:
                    ngay_kc = str(nkc)
        
        # Vật tư
        vt_tcty = 0
        vt_dv = 0
        if not vt_df.empty:
            vt_tcty = _safe_int(vt_df.iloc[0].get('TCty cấp', 0))
            vt_dv = _safe_int(vt_df.iloc[0].get('ĐV cấp', 0))
        
        # KL thực hiện - tổng GT thực hiện từ tất cả HĐ
        gt_thuc_hien_hd = 0
        tong_gt_hd = 0
        if not hd_df.empty:
            gt_thuc_hien_hd = hd_df['Giá trị thực hiện HĐ'].apply(_safe_int).sum()
            tong_gt_hd = hd_df['Giá trị HĐ'].apply(_safe_int).sum()
        
        ty_le_th = round(gt_thuc_hien_hd / tong_gt_hd * 100, 1) if tong_gt_hd > 0 else 0
        
        # KL thanh toán - from PM_092
        ty_le_tt = round(thuc_hien / tong_gt_hd * 100, 1) if tong_gt_hd > 0 else 0
        
        # Nghiệm thu
        ngay_nt = ''
        gt_qt = 0
        ghi_chu = ''
        if not nt_df.empty:
            r = nt_df.iloc[0]
            nnt = r.get('Ngày nghiệm thu CT', '')
            if pd.notna(nnt):
                try:
                    if isinstance(nnt, pd.Timestamp):
                        ngay_nt = nnt.strftime('%d/%m/%Y')
                    else:
                        ngay_nt = str(nnt)
                except:
                    ngay_nt = str(nnt)
            gt_qt = _safe_int(r.get('Giá trị quyết toán CT', 0))
            ghi_chu = _safe_val(r.get('Ghi chú', ''))
        
        # Build row (values in đồng for report)
        hm_row = {
            '_category': CAT_MAP.get(row.get('Tên hạng mục', ''), 'Sửa chữa khác'),
            'STT': stt_counter,
            'Mã công trình': ma_ct,
            'Đơn vị': 'Công ty Điện lực Vũng Tàu',
            'Tên công trình': ten_ct,
            'Nội dung': _safe_val(row.get('Nội dung SCL', '')),
            'Mã TS': '',
            'Tiến độ thực hiện': _safe_val(row.get('Tiến độ', '')),
            'Giá trị khái toán': _to_dong(khai_toan),
            'PAKT-DT Số QĐ ngày duyệt': pakt_soqd,
            'PAKT-DT Giá trị DT': _to_dong(pakt_gt),
            'Giá trị vốn KH': _to_dong(ke_hoach),
            'KH ĐT Gói XL Số QĐ': kh_xl_soqd,
            'KH ĐT Gói XL GT': _to_dong(kh_xl_gt),
            'KH ĐT Gói TB Số QĐ': kh_tb_soqd,
            'KH ĐT Gói TB GT': _to_dong(kh_tb_gt),
            'KQ ĐT Gói XL Số QĐ': kq_xl_soqd,
            'KQ ĐT Gói XL GT': _to_dong(kq_xl_gt),
            'KQ ĐT Gói XL Số HĐ': hd_xl_so,
            'KQ ĐT Gói XL GT HĐ': _to_dong(hd_xl_gt),
            'KQ ĐT Gói TB Số QĐ': kq_tb_soqd,
            'KQ ĐT Gói TB GT': _to_dong(kq_tb_gt),
            'KQ ĐT Gói TB Số HĐ': hd_tb_so,
            'KQ ĐT Gói TB GT HĐ': _to_dong(hd_tb_gt),
            'Đơn vị thi công': don_vi_tc,
            'Ngày khởi công': ngay_kc,
            'VT TCty cấp': _to_dong(vt_tcty),
            'VT ĐV cấp': _to_dong(vt_dv),
            'KL thực hiện GT': _to_dong(gt_thuc_hien_hd),
            'KL thực hiện %': ty_le_th,
            'KL thanh toán GT': _to_dong(thuc_hien),
            'KL thanh toán %': ty_le_tt,
            'Ngày nghiệm thu': ngay_nt,
            'Giá trị quyết toán': _to_dong(gt_qt),
            'Ghi chú': ghi_chu,
        }
        rows.append(hm_row)
    
    if not rows:
        return None
        
    final_rows = []
    ORDER = ['Lưới và trạm điện', 'Thiết bị', 'Kiến trúc', 'Công xa', 'Sửa chữa khác']
    numeric_cols = [
        'Giá trị khái toán', 'PAKT-DT Giá trị DT', 'Giá trị vốn KH',
        'KH ĐT Gói XL GT', 'KH ĐT Gói TB GT', 
        'KQ ĐT Gói XL GT', 'KQ ĐT Gói XL GT HĐ',
        'KQ ĐT Gói TB GT', 'KQ ĐT Gói TB GT HĐ',
        'VT TCty cấp', 'VT ĐV cấp',
        'KL thực hiện GT', 'KL thanh toán GT', 'Giá trị quyết toán'
    ]
    
    total_counts = 0
    grand_totals = {c: 0 for c in numeric_cols}
    
    for cat in ORDER:
        cat_items = [r for r in rows if r.get('_category') == cat]
        if not cat_items:
            continue
            
        cat_counts = len(cat_items)
        total_counts += cat_counts
        
        # Tính tổng của nhóm cho các cột số
        cat_totals = {c: sum(_safe_int(r.get(c, 0)) for r in cat_items) for c in numeric_cols}
        for c in numeric_cols:
            grand_totals[c] += cat_totals[c]
            
        # Thêm dòng header của nhóm
        header_row = {c: '' for c in HANGMUC_COLS}
        header_row['Mã công trình'] = cat
        header_row['Tên công trình'] = str(cat_counts)
        for c in numeric_cols:
            header_row[c] = cat_totals[c]
            
        header_row['_is_header'] = True
        final_rows.append(header_row)
        
        # Thêm các công trình, đánh lại STT từ 1
        stt = 1
        for r in cat_items:
            r['STT'] = stt
            stt += 1
            final_rows.append(r)
            
    # Thêm dòng TỔNG CỘNG ở cuối
    if final_rows:
        total_row = {c: '' for c in HANGMUC_COLS}
        total_row['Mã công trình'] = 'TỔNG CỘNG'
        total_row['Tên công trình'] = str(total_counts)
        for c in numeric_cols:
            total_row[c] = grand_totals[c]
        total_row['_is_total'] = True
        final_rows.append(total_row)
    
    result_df = pd.DataFrame(final_rows, columns=HANGMUC_COLS)
    
    import openpyxl
    from copy import copy
    from openpyxl.styles import Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    
    template_path = os.path.join(BASE_DIR, 'MauHangMuc.xlsx')
    if os.path.exists(template_path):
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        # Đọc style từ dòng 5 (dòng dữ liệu chuẩn trong file template)
        style_row = 5
        styles = []
        for col in range(1, len(HANGMUC_COLS) + 1):
            cell = ws.cell(row=style_row, column=col)
            styles.append({
                'font': copy(cell.font) if cell.font else None,
                'fill': copy(cell.fill) if cell.fill else None,
                'alignment': copy(cell.alignment) if cell.alignment else None
            })
        
        # Khởi tạo định dạng border mỏng cho tất cả ô
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Xóa các dòng dữ liệu mẫu (từ dòng 4 trở đi)
        if ws.max_row >= 4:
            ws.delete_rows(4, ws.max_row - 3)
            
        # Biến lưu trữ độ rộng các cột
        col_widths = {c: 0 for c in range(1, len(HANGMUC_COLS) + 1)}
        
        # Đọc trước độ rộng của header (từ dòng 1 đến 3)
        for c in range(1, len(HANGMUC_COLS) + 1):
            max_w = 0
            for r in range(1, 4):
                cv = ws.cell(row=r, column=c).value
                if cv:
                    lines = str(cv).split('\n')
                    max_w = max(max_w, max(len(l) for l in lines))
            col_widths[c] = max_w
        
        # Ghi dữ liệu mới vào
        start_row = 4
        for r_idx, r_dict in enumerate(final_rows, start=start_row):
            is_bold_row = r_dict.get('_is_header') or r_dict.get('_is_total')
            for c_idx, col_name in enumerate(HANGMUC_COLS, start=1):
                val = r_dict.get(col_name, '')
                c = ws.cell(row=r_idx, column=c_idx, value=val)
                
                # Apply style từ template
                s = styles[c_idx - 1]
                if s['font']: 
                    new_f = copy(s['font'])
                    if is_bold_row:
                        new_f.bold = True
                    c.font = new_f
                elif is_bold_row:
                    c.font = openpyxl.styles.Font(bold=True)
                    
                if s['fill']: c.fill = copy(s['fill'])
                
                # 1. Đóng khung toàn bộ ô
                c.border = thin_border
                
                # 2. Định dạng số
                if isinstance(val, (int, float)) and val != '':
                    if '%' in col_name or 'Tỉ lệ' in col_name or 'tỷ lệ' in col_name.lower() or 'Tỷ lệ' in col_name:
                        c.number_format = '0.00'
                    else:
                        c.number_format = '#,##0'
                
                # 3. Canh lề và wrap_text
                if col_name == 'Nội dung':
                    c.alignment = Alignment(wrap_text=True, vertical='top')
                elif s['alignment']:
                    c.alignment = copy(s['alignment'])
                    
                # 4. Tính toán độ rộng để dãn cột
                if val is not None and val != '':
                    lines = str(val).split('\n')
                    max_w = max(len(l) for l in lines)
                    # Cộng thêm 1 chút không gian cho định dạng số có dấu phẩy
                    if isinstance(val, (int, float)):
                        max_w += 3 
                    if max_w > col_widths[c_idx]:
                        col_widths[c_idx] = max_w
                        
        # Áp dụng dãn cột
        for c_idx, max_w in col_widths.items():
            width = max_w + 3
            # Giới hạn cột Nội dung tối đa là 70
            if HANGMUC_COLS[c_idx - 1] == 'Nội dung':
                width = min(width, 70)
            # Giới hạn tối thiểu
            width = max(width, 8)
            ws.column_dimensions[get_column_letter(c_idx)].width = width
        
        wb.save(HANGMUC_OUTPUT)
    else:
        # Fallback nếu không có file template
        result_df.to_excel(HANGMUC_OUTPUT, index=False)
        
    return result_df
